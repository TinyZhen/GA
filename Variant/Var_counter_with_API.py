
from tkinter import *
from tkinter import ttk,filedialog,messagebox
import pickle,os
from pathlib import Path
import pandas as pd 
import openpyxl
import requests
import re
from io import StringIO 
from dateutil.relativedelta import relativedelta
from tkcalendar import DateEntry


end_date = None
start_date = None
base_url = 'https://lapis.cov-spectrum.org/open/v1/sample/' 
mutation_url = None
lineage_url = 'aa-mutations?pangoLineage='
format_fild = '&dataFormat=csv'
search_words = ['Sequences','Unique Sequence']
workbook = None
sheet = None
counter =0

def inputfile():
    raw_input.set(filedialog.askdirectory())
    entry1.xview_moveto(1)

def saveresult():
    result_file.set(filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Excel files", "*.xlsx"),("all files",
                                                        "*.*")]))
    entry4.xview_moveto(1)

def clear():
    raw_input.set("")
    result_file.set("")
    counter=0
    cal.delete(0,'end')
    log.delete(1.0, END) 

def close():
    pickle.dump(raw_input.get(),open("pref.dat", "wb"))
    root.destroy()

def logprint(text):
    log.insert(END, text + '\n')

def handle_df(df,sit):
    global counter
    # print(sit)
    total_count = df['Count'].sum()
    # iterate through sequences
    for indexs, row in df.iterrows():
        sequence = row['Sequences']
        cell = sheet[f'F{counter}']
        cell.value = sequence
        count = row['Count']
        flag = 0
        aamutation_list = parse(sequence)
        api = generateURL(aamutation_list,mutation_url,format_fild)
        response = API_call(api)
        if response.status_code == 200:
        # Parse the CSV data from the response content
            csv_data = response.text
            
            # Use StringIO to convert the CSV data into a file-like object
            csv_file = StringIO(csv_data)
            
            # Create a CSV reader and iterate through the rows
            lineage_df = pd.read_csv(csv_file)

            if not lineage_df.empty:
                lineage_sorted_df = lineage_df.sort_values(by='count', ascending=False)
                total_count2 = lineage_df['count'].sum()
                lineage_sorted_df['Abundance'] = lineage_sorted_df['count']/total_count2
                for index, rows in lineage_sorted_df.iterrows():
                    # only checking the abundance of lineage that is greater than 5%
                    print(rows["count"])
                    if rows['Abundance'] > 0.05:
                        api = generateURL(rows['pangoLineage'],lineage_url,format_fild)
                        response = API_call(api)
                        if response.status_code == 200:
                            # Parse the CSV data from the response content
                            csv_data = response.text
                            
                            # Use StringIO to convert the CSV data into a file-like object
                            csv_file = StringIO(csv_data)
                            
                            # Create a CSV reader and iterate through the rows
                            mutation_df = pd.read_csv(csv_file)
                            result_list = check_mutation(aamutation_list,mutation_df)
                        if(not result_list):
                            # print(f'Sequence:{indexs+1} {rows["pangoLineage"]} Percentage:{count/total_count*100:.1f}')
                            update_excel(rows["pangoLineage"],count/total_count*100,'G',counter)
                            break
                        else:
                            # print(f'Sequence:{indexs+1} Possible lineage: {rows["pangoLineage"]} Percentatge:{count/total_count*100:.1f}')
                            if not flag:
                                update_excel("Other",count/total_count*100,'H',counter)
                                flag = 1
                            update_excel(rows["pangoLineage"],result_list,counter)
            else:
                # print('not found')
                # print(f'Sequence:{indexs+1} {sequence} Percentatge:{count/total_count*100:.1f}')
                update_excel("Other",count/total_count*100,'H',counter)                
        else:
            # print('not found')
            # print(f'Sequence:{indexs+1} {sequence} Percentatge:{count/total_count*100:.1f}')
            update_excel("Other",count/total_count*100,'H',counter)
        counter += 2

def check_mutation(seq, m_df):
    sorted_df = sort_df(m_df)
    column_items = set(sorted_df['mutation'])
    seq_items = set(seq)
    mismatches = column_items.symmetric_difference(seq_items)
    if mismatches:
        # Check if there are extra items in the column or seq
        extra_in_column = column_items - seq_items
        extra_in_seq = seq_items - column_items

        # if extra_in_column:
        #     print(f"Missing: {extra_in_column}")
        # if extra_in_seq:
        #     print(f"Extra: {extra_in_seq}")
        if extra_in_column and extra_in_seq:
            return [f"Missing: {extra_in_column}",f"Extra: {extra_in_seq}"]
        elif extra_in_column:
            return[f"Missing: {extra_in_column}"]
        elif extra_in_seq:
            return[f"Extra: {extra_in_seq}"]
        else:
            return [f"Missmatched: {mismatches}"]

    else:
        return []

def find_col(row):
    for column in range(8, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            if cell.value is not None:
                last_column = openpyxl.utils.get_column_letter(column)
    
    next_column = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(last_column) + 1)
    return next_column

def update_excel(*args, **kwargs):
    global sheet
    if len(args) == 1:
        return args[0]
    elif len(args) == 4:
        c1 = sheet[f'{args[2]}{args[3]}']
        c1.value = args[0]
        c2 = sheet[f'{args[2]}{args[3] + 1}']  
        c2.value = args[1]
    elif len(args) == 3:
        col = find_col(args[2])
        c1 = sheet[f'{col}{args[2]}'] 
        c1.value = args[0]
        c2 = sheet[f'{col}{args[2]+1}']  
        c2.value = f'{args[1]}'

def create_excel():
    global workbook, sheet

    data = pd.DataFrame()
    data['Label'] = None
    data['Code'] = None
    data['Date'] = None
    data['Algorithm'] = None
    data['Number of Reads'] = None
    data['Sequence'] = None
    data['Match'] = None
    data['Unkown'] = None
    try:
        data.to_excel(result_file.get(),index=False)
        
    except:
        logprint("- Error writing output file!")
        messagebox.showerror('Error', 'Error writing output file!') 
        return 1
    logprint("Creating Excel file...")
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(result_file.get())

        # Select the active (current) sheet
        sheet = workbook.active        
    except:
        logprint("- Error load output file!")
        return 1

def addto_excel(sit,df,file):
    global sheet,workbook
    pattern = r'^EC_.*?15(.*?)(\d+)'
    match = re.search(pattern,sit)
    start_col = 'A'
    if match:
        code = match.group(1)
        date = match.group(2)
        if file.lower().endswith("chim_rm.tsv"):
            algorithm = 'Chimeras_Removed'
        elif file.lower().endswith("covar_deconv.tsv"):
            algorithm = 'Covar_Deconv'
        total_count = df['Count'].sum()
        num_row = len(df['Sequences'])*2
        data=[sit,code,date,algorithm,total_count]
        merge_cell(num_row)
        # add data to the sheet
        for item in data:
            cell = sheet[f'{start_col}{sheet.max_row}']  # Get the cell in the specified column and row
            top_left_cell = None
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = sheet.cell(row=merged_range.bounds[1], column=merged_range.bounds[0])
                    break
            top_left_cell.value = item
            # add wrap text
            start_col = chr(ord(start_col) + 1)

def merge_cell(n):
    global workbook, sheet

    # Calculate the starting row for merging (1 row below the current max row)
    start_row = sheet.max_row + 1

    # Define the merge range based on the starting row and the number of rows
    end_row = start_row + n - 1
    for col in range(1, 6):
        column_letter = openpyxl.utils.get_column_letter(col)
        merge_range = f'{column_letter}{start_row}:{column_letter}{end_row}'
        sheet.merge_cells(merge_range)

def sort_df(m_df):
    pattern = r'^S:[A-Z]?(\d+)[A-Z]?$'
    matches = []

    # Loop through each row in the DataFrame
    for index, row in m_df.iterrows():
        mutation = row['mutation']
        
        # find the pattern in the mutation string
        match = re.search(pattern, mutation)
        
        # Check if a match was found and if the numeric part is between 300 and 505
        if match:
            numeric_part = int(match.group(1))
            if 300 <= numeric_part <= 505:
                matches.append(row)
    matched_df = pd.DataFrame(matches)
    return matched_df
        
def generateURL(lst,url,format):
    api = f'{base_url}{url}'
    if(isinstance(lst,list)):
        item_str = ','.join(lst)
    else:
        item_str = lst
    api = f'{api}{item_str}{format}'
    return api

def parse(str):
    matches = re.findall(r'\((.*?)\)', str)
    matches_with_prefix = [f'S:{match}' for match in matches]
    return matches_with_prefix

def API_call(url):
    response = requests.get(url)
    return response

def result():
    global workbook,counter,end_date,start_date,mutation_url
    end_date = cal.get_date()
    start_date = end_date - relativedelta(months=6)
    mutation_url = f'aggregated?&dateFrom={start_date}&dateTo={end_date}&fields=pangoLineage&aaMutations='
    inpath = Path(raw_input.get())
    counter = 2
    logprint("Adding data to file...")
    logprint("Loading...")
    try:
        ### GET LIST OF VALID INPUT FILES ###
        tsv_files = [
            os.path.join(root, file)
            for root, _, files in os.walk(inpath)
            for file in files
            if (
                file.lower().endswith("chim_rm.tsv") or
                file.lower().endswith("covar_deconv.tsv")

            ) and "Collected" not in file
            # if file.lower().endswith(".tsv")

        ]
        # for tsv_file in tsv_files:
            # print("Matching TSV File:", tsv_file)
    except:
        logprint("- Error opening directory!")
        messagebox.showerror('Error', 'Error opening directory!')
        return 1

    if not tsv_files:
        logprint("- No files were found!")
        messagebox.showwarning('Warning', 'No input files found!')
        return 0
    
    tsv_files.sort()
    create_excel()
    for file in tsv_files:      
        try:
            df = pd.read_csv(inpath / file,sep='\t')
            sit = df.columns[0]
            df = pd.read_csv(inpath / file,sep='\t',header=1)
        except:
            logprint("- Error opening %s!" % file)
            messagebox.showerror('Error', 'Error opening %s!' % file)
            return 1
        
        for index, row in df.iterrows():
            sequence = row['Sequences']
            pattern = r'\S*fs\S*'
            matches = re.findall(pattern, sequence)
            if matches:
                df = df.drop(index, axis=0)

        # add output to sheeet
        addto_excel(sit,df,file)
        handle_df(df,sit)
        print("Loading...")
    workbook.save(result_file.get())
    logprint("Done writting to file...")

if __name__ == '__main__':
    root = Tk()
    root.title("Variant spotter")
    raw_input = StringVar()
    result_file = StringVar()


    try:
        raw_input.set(pickle.load(open( "pref.dat", "rb" )))
    except:
        pass

    frm = ttk.Frame(root, padding=10)
    frm.pack(side=LEFT)

    row1 = ttk.Frame(frm)
    cal=DateEntry(row1,width=12, background='white',
                    foreground='black', borderwidth=2)
    cal.pack(side=RIGHT, padx=5)
    cal.config(state="readonly")

    label = ttk.Label(row1, text="Select a date:")
    label.pack(side=LEFT, padx=5)
    row1.pack(side=TOP, padx=5, pady=5)

    row2 = ttk.Frame(frm)
    ttk.Button(row2, width=25, text="Load Raw data:", command=inputfile).pack(side=LEFT,padx=5)
    entry1 = ttk.Entry(row2,width=40,textvariable=raw_input)
    entry1.config(state="readonly")
    row2.pack(side=TOP, padx=5, pady=5)
    entry1.pack(side=RIGHT, expand=YES, fill=X)
    entry1.xview_moveto(1)

    row4 = ttk.Frame(frm)
    ttk.Button(row4, width=25, text="Save results as:", command=saveresult).pack(side=LEFT,padx=5)
    entry4 = ttk.Entry(row4,width=40,textvariable=result_file)
    entry4.config(state="readonly")
    row4.pack(side=TOP, padx=5, pady=5)
    entry4.pack(side=RIGHT, expand=YES, fill=X)

    buttonrow2 = ttk.Frame(frm)
    ttk.Button(buttonrow2, text="Run", command=result).pack(side=LEFT,padx=15)
    ttk.Button(buttonrow2, text="Clear", command=clear).pack(side=LEFT,padx=15)
    ttk.Button(buttonrow2, text="Close", command=close).pack(side=LEFT,padx=15)
    buttonrow2.pack(side=BOTTOM,pady=5)

    log = Text(height=20,width=50)
    log.pack(side=RIGHT, padx=10,pady=10)

    root.protocol("WM_DELETE_WINDOW", close)

    root.mainloop()

