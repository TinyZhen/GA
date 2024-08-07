
from tkinter import *
from tkinter import ttk,filedialog,messagebox
import pickle,os,sys,csv
from pathlib import Path
import pandas as pd 
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import xlsxwriter
import os
import math
from numbers import Number


search_word = "Sample Name"
sheet_name = "Results"
starting_word = "Well"
filter_word = ["Sample Name", "Ct Mean","Ct SD"]
sample =[]
machine1 = { "slop": -3.46,
             "y-int": 39.1,
             "efficiency": 0.9454
            }
machine2 = { "slop": -3.38,
             "y-int": 38.4,
             "efficiency": 0.9763
            }
wb = openpyxl.Workbook()
red_fill = PatternFill(patternType='solid', fgColor= '00FF0000')

def inputmaster():
    master_input.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"),("all files",
                                                        "*.*")]))
    
def inputfile():
    raw_input.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"),("all files",
                                                        "*.*")]))
    entry1.xview_moveto(1)
    
def saveresult():
    ofname.set(filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Excel files", "*.xlsx"),("all files",
                                                        "*.*")]))
    entry4.xview_moveto(1)

def close():
    pickle.dump(raw_input.get(),open("pref.dat", "wb"))
    root.destroy()

def logprint(text):
    log.insert(END, text + '\n')

def mapping():
    global buttonrow, machine
    try:
        inpath = Path(raw_input.get())
        df = pd.DataFrame()
        df = pd.read_excel(inpath, sheet_name=sheet_name)
    except:
        logprint("- Error reading input file!")
        messagebox.showerror('Error', 'Error reading input file!') # TODO show output file name
        return 1
    
    for index, row in df.iterrows():
        if row.str.contains("Instrument Type").any():
            machine_row = row
        if row.str.contains(starting_word).any():
            start_index = index
            break
    
    if start_index is not None:
        # Create a new DataFrame starting from the row containing the start_word
        new_df = pd.DataFrame()
        new_df = df.iloc[start_index:]
        new_df.reset_index(drop=True, inplace=True)
        new_df.columns = new_df.iloc[0]
        new_df = new_df.iloc[1:]
        new_df.reset_index(drop=True, inplace=True)
    else:
        print(f"Starting cell with the word '{starting_word}' not found in the DataFrame.")

    if machine_row is not None:
        machine_val = machine_row[1]  # Assuming "Instrument Name" is in the second column
        if "3" in machine_val:
            machine = machine1
        else:
            machine = machine2
        machine_str.set(f"{machine_val}")
        
    return new_df

def findneg(df,wb,sheet):
    for row, value in enumerate(df[search_word],start=2):
         if isinstance(value, str) and len(value) > 0:
            if value.startswith("NEG"):
                ct_mean = df.loc[row - 2, 'Ct Mean']
                if not pd.isna(ct_mean) or ct_mean < 38:
                    cell = sheet.cell(row,column=2)
                    cell.fill = red_fill
            elif value.startswith("NTC"):
                ct_mean = df.loc[row - 2, 'Ct Mean']
                if not pd.isna(ct_mean):
                    cell = sheet.cell(row,column=2)
                    cell.fill = red_fill
            elif value.startswith("EXT"):
                ct_mean = df.loc[row - 2, 'Ct Mean']
                if not pd.isna(ct_mean) or ct_mean < 38:
                    cell = sheet.cell(row,column=2)
                    cell.fill = red_fill

    wb.save(ofname.get())

def checkcal(df,wb,sheet):
    for row, value in enumerate(df[search_word],start=2):
        if isinstance(value, str) and len(value) > 0:
            if value.startswith("Cal"):
                ct_mean = df.loc[row - 2, 'Ct Mean']
                if ct_mean < 27 or ct_mean > 28.5 :
                    cell = sheet.cell(row,column=2)
                    cell.fill = red_fill
    wb.save(ofname.get())

def checkmean_and_sd(df,wb, sheet):

    # Iterate through the 'Ct Mean' column and fill with red if > 31
    for row_number, ct_mean_value in enumerate(df['Ct Mean'], start=2):  # Start at row 2 (assuming header is in row 1)
        if ct_mean_value > 31:
            cell = sheet.cell(row_number,column=2)
            cell.fill = red_fill

    for row_number, ct_sd_value in enumerate(df['Ct SD'], start=2):  # Start at row 2 (assuming header is in row 1)
        if ct_sd_value > 0.2:
            cell = sheet.cell(row_number,column=3)
            cell.fill = red_fill
    # Save the Excel file with modified formatting
    wb.save(ofname.get())

def checkresult(df):
    try:
        wb = openpyxl.load_workbook(ofname.get())

        # Access the sheet
        sheet = wb['Sheet1']
    except:
        logprint("cant open the file")
    checkmean_and_sd(df,wb,sheet)
    findneg(df,wb,sheet)
    checkcal(df,wb,sheet)

def updatedropdown():
    global clicked2, buttonrow3, drop2
    drop2['menu'].delete(0,'end')
    clicked2.set(sheet_names[0])
    for name in sheet_names:
        drop2['menu'].add_command(label=name, command=lambda name=name: clicked2.set(name))

def     loadconcerntrate():
    global sheet_names, wb
    sheet_names = []
    try:
        wb = openpyxl.load_workbook(master_input.get())
        sheet_names = wb.sheetnames

    except:
        logprint("cant open the file")
    if not sheet_names:
        logprint("No sheets found")
        return
    else:
        updatedropdown()
    logprint("- Loading Master sheets done !")

def add_concentrate_dilution(df):
    global wb
    try:
        df1 = df
        sheet = wb[clicked2.get()]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Convert the list of lists to a pandas DataFrame
        df2 = pd.DataFrame(data)
        df2.columns = df2.iloc[0]
        df2 = df2[1:]
        # df2 = pd.read_excel(master_input.get(),clicked2.get())
        for index, row in df1.iterrows():   
            master_id = row[search_word]
            if not pd.isna(master_id):
                matching_row_df2 = df2[df2['[Sample ID]'].astype(str).str.contains(str(master_id))]
            if not matching_row_df2.empty:
        # Extract the value from the matching row in df2
                for indexs in matching_row_df2.index:
                    concern = matching_row_df2.loc[indexs, '[Final Concentrate Volume (mL)]']
                    dilution = matching_row_df2.loc[indexs, '[Dilution factor]']
                    df1.at[index, "Concentrate Volume (mL)"] = concern
                    df1.at[index,"Dilution Factor"] = dilution
    except:
        logprint("- Error add concentration !")
        return 1
    return df1

def outputConstrain(sample):
    # make an empty output constrain
    try:
        # write to the file
        sample_df = pd.DataFrame()
        sample_df = pd.DataFrame({search_word: sample,
                                    "Dilution Factor": [np.nan] * len(sample),
                                    "Concentrate Volume (mL)": [np.nan] * len(sample)
                                    
        })
    except:
        logprint("- Error writing constrain file!")
        messagebox.showerror('Error', 'Error writing constrain file!') 
        return 1
    df = add_concentrate_dilution(sample_df)
    return df

def output_to_Excel(data,name):
    try:

        data.to_excel(name,index=False)
        
    except:
        logprint("- Error writing output file!")
        messagebox.showerror('Error', 'Error writing output file!') 
        return 1
    
def extractConstrain():
    df = pd.DataFrame()
    df = mapping()
    # find the unique Sample ID and extract the column
    data = []
    sample = []
    data = df[search_word].unique()
    for x in data:
        # filter the control sets
        if (len(str(x)) >= 9):
            sample.append(x)
    return sample

def runConstrain():
        df = pd.DataFrame()
        df = extractConstrain()
        df1 = outputConstrain(df)
        return df1


def calculatePMMoV(ct_mean,concentrate,dilution):
    global machine
    pmmov1 = 10 ** ((ct_mean-machine["y-int"])/machine["slop"])
    pmmov2 = pmmov1/5 * dilution
    pmmov3 = pmmov2 *80
    pmmov4 = pmmov3/0.2*concentrate
    return pmmov4

def result():
    # read two excel
    df1 = mapping()
    df2 = runConstrain()

    # merge two excel
    df3 = df1[filter_word].merge(df2,on = search_word, how = "left")

    #filter the excel
    df3['Sample Name'] = df3['Sample Name'].str.strip()
    df3_filter= df3.drop_duplicates()
    df3_filter = df3_filter.reset_index(drop=True)
    df3_filter["Calibrator Ct Mean"] = None
    df3_filter["PMMoV (gc/100 mL Sewage)"] = None

    for index,row in df3_filter.iterrows():
        ct_mean = row["Ct Mean"]
        concentrate = row["Concentrate Volume (mL)"]
        dilution = row["Dilution Factor"]

        if (isinstance(dilution, float) and isinstance(concentrate, Number)):
            if not (math.isnan(dilution) or math.isnan(concentrate)):
                dilution = int(dilution)  # Convert 'dilution' to an integer
            
        else:
            if not isinstance(dilution, float):
                dilution = 0
            if not isinstance(concentrate, float):
                concentrate = 0
        pmmov = calculatePMMoV(ct_mean,concentrate,dilution)
        df3_filter.loc[index, "PMMoV (gc/100 mL Sewage)"] = "{:.2e}".format(pmmov)

    
    for row, value in enumerate(df3_filter[search_word],start=2):
        if isinstance(value, str) and len(value) > 0:
            if value.startswith("Cal"):
                ct_mean = df3_filter.loc[row - 2, 'Ct Mean']
                df3_filter["Calibrator Ct Mean"] = ct_mean
    output_to_Excel(df3_filter,ofname.get())
    checkresult(df3_filter)
    logprint("- Outputting result file done !")
    try:
        os.system(f'start excel "{ofname.get()}"')
    except Exception as e:
        print(f"An error occurred: {e}")


def clear():
    global sheet_names,drop2
    raw_input.set("")
    constrain_input.set("")
    ofname.set("")
    constrain_output.set("")
    master_input.set("")
    sheet_names = []
    clicked2.set(' ')
    drop2['menu'].delete(0,END)
    machine_str.set("")
    log.delete(1.0, END) 
    
if __name__ == '__main__':
    root = Tk()
    root.title("PMMoV Calculation")
    raw_input = StringVar()
    constrain_input = StringVar()
    ofname = StringVar()
    constrain_output = StringVar()
    master_input = StringVar()
    clicked2 = StringVar()
    machine_str = StringVar()
    machine = {}
    try:
        raw_input.set(pickle.load(open( "pref.dat", "rb" )))
    except:
        pass

    frm = ttk.Frame(root, padding=10)
    frm.pack(side=LEFT)

    row5 = ttk.Frame(frm)
    ttk.Button(row5, width=25, text="Load Master sheet:", command=inputmaster).pack(side=LEFT,padx=5)
    entry5 = ttk.Entry(row5,width=40,textvariable=master_input)
    entry5.config(state="readonly")
    row5.pack(side=TOP, padx=5, pady=5)
    entry5.pack(side=RIGHT, expand=YES, fill=X)
    entry5.xview_moveto(1)

    buttonrow3 = ttk.Frame(frm)
    ttk.Button(buttonrow3, text="Load", command=loadconcerntrate).pack(side=LEFT,padx=15)
    drop2 = OptionMenu(buttonrow3,clicked2, [])
    drop2.pack(side=LEFT,pady=5)
    buttonrow3.pack(side=TOP,pady=5)

    row1 = ttk.Frame(frm)
    ttk.Button(row1, width=25, text="Load Raw data:", command=inputfile).pack(side=LEFT,padx=5)
    entry1 = ttk.Entry(row1,width=40,textvariable=raw_input)
    entry1.config(state="readonly")
    row1.pack(side=TOP, padx=5, pady=5)
    entry1.pack(side=RIGHT, expand=YES, fill=X)

    row4 = ttk.Frame(frm)
    ttk.Button(row4, width=25, text="Save results as:", command=saveresult).pack(side=LEFT,padx=5)
    entry4 = ttk.Entry(row4,width=40,textvariable=ofname)
    entry4.config(state="readonly")
    row4.pack(side=TOP, padx=5, pady=5)
    entry4.pack(side=RIGHT, expand=YES, fill=X)

    buttonrow2 = ttk.Frame(frm)
    ttk.Button(buttonrow2, text="Run", command=result).pack(side=LEFT,padx=15)
    ttk.Button(buttonrow2, text="Clear", command=clear).pack(side=LEFT,padx=15)
    ttk.Button(buttonrow2, text="Close", command=close).pack(side=LEFT,padx=15)
    buttonrow2.pack(side=BOTTOM,pady=5)


    
    log = Text(height=20,width=50)
    log.pack(side=RIGHT, padx=10,pady=10)

    root.protocol("WM_DELETE_WINDOW", close)

    root.mainloop()
