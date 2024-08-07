
from tkinter import *
from tkinter import ttk,filedialog,messagebox
import pickle,os,sys,csv
from pathlib import Path
import pandas as pd 
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import xlsxwriter
import os
import math
from numbers import Number

class PMMoVCalc:

    search_word = "Sample Name"
    sheet_name = "Results"
    starting_word = "Well"
    filter_word = ["Sample Name", "Ct Mean","Ct SD"]
    sample =[]
    machine1 = { "slop": -3.46,
                "y-int": 39.1,
                "efficiency": 0.9454
                }
    machine2 = { "slop": -3.38,
                "y-int": 38.4,
                "efficiency": 0.9724
                }
    wb = openpyxl.Workbook()
    red_fill = PatternFill(patternType='solid', fgColor= '00FF0000')

    LOWER =27
    UPPER =28.5

    def __init__(self, root):
        self.root = root
        self.root.title("PMMoV Calculation")

        self.create_widgets()




    def create_widgets(self):
        global entry1,entry4,buttonrow3,drop2
        self.raw_input = StringVar()
        self.constrain_input = StringVar()
        self.ofname = StringVar()
        self.constrain_output = StringVar()
        self.master_input = StringVar()
        self.clicked2 = StringVar()
        self.machine_str = StringVar()
        self.machine1_slop = StringVar()
        self.machine2_slop = StringVar()
        self.machine1_int = StringVar()
        self.machine2_int = StringVar()
        self.lower_cal = StringVar()
        self.upper_cal = StringVar()

        self.machine1_slop.set(self.machine1["slop"])
        self.machine2_slop.set(self.machine2["slop"])
        self.machine1_int.set(self.machine1["y-int"])
        self.machine2_int.set(self.machine2["y-int"])
        self.lower_cal.set(self.LOWER)
        self.upper_cal.set(self.UPPER)
        try:
            self.raw_input.set(pickle.load(open( "pref.dat", "rb" )))
        except:
            pass

        frm = ttk.Frame(self.root, padding=10)
        frm.pack(side=LEFT)

        row2 = ttk.Frame(frm)
        ttk.Button(row2, width=25, text="Setting", command=self.setting).pack(side=LEFT, padx=5)
        row2.pack(side=TOP, padx=5, pady=5)

        row5 = ttk.Frame(frm)
        ttk.Button(row5, width=25, text="Load Master sheet:", command=self.inputmaster).pack(side=LEFT,padx=5)
        entry5 = ttk.Entry(row5,width=40,textvariable=self.master_input)
        entry5.config(state="readonly")
        row5.pack(side=TOP, padx=5, pady=5)
        entry5.pack(side=RIGHT, expand=YES, fill=X)
        entry5.xview_moveto(1)

        buttonrow3 = ttk.Frame(frm)
        ttk.Button(buttonrow3, text="Load", command=self.loadconcerntrate).pack(side=LEFT,padx=15)
        drop2 = OptionMenu(buttonrow3,self.clicked2, [])
        drop2.pack(side=LEFT,pady=5)
        buttonrow3.pack(side=TOP,pady=5)

        row1 = ttk.Frame(frm)
        ttk.Button(row1, width=25, text="Load Raw data:", command=self.inputfile).pack(side=LEFT,padx=5)
        entry1 = ttk.Entry(row1,width=40,textvariable=self.raw_input)
        entry1.config(state="readonly")
        row1.pack(side=TOP, padx=5, pady=5)
        entry1.pack(side=RIGHT, expand=YES, fill=X)

        row4 = ttk.Frame(frm)
        ttk.Button(row4, width=25, text="Save results as:", command=self.saveresult).pack(side=LEFT,padx=5)
        entry4 = ttk.Entry(row4,width=40,textvariable=self.ofname)
        entry4.config(state="readonly")
        row4.pack(side=TOP, padx=5, pady=5)
        entry4.pack(side=RIGHT, expand=YES, fill=X)

        buttonrow2 = ttk.Frame(frm)
        ttk.Button(buttonrow2, text="Run", command=self.result).pack(side=LEFT,padx=15)
        ttk.Button(buttonrow2, text="Clear", command=self.clear).pack(side=LEFT,padx=15)
        ttk.Button(buttonrow2, text="Close", command=self.close).pack(side=LEFT,padx=15)
        buttonrow2.pack(side=BOTTOM,pady=5)
        
        self.log = Text(height=20,width=50)
        self.log.pack(side=RIGHT, padx=10,pady=10)

        self.root.protocol("WM_DELETE_WINDOW", self.close)

    def inputmaster(self):
        self.master_input.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"),("all files",
                                                            "*.*")]))
        
    def inputfile(self):
        self.raw_input.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"),("all files",
                                                            "*.*")]))
        entry1.xview_moveto(1)
        
    def saveresult(self):
        self.ofname.set(filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Excel files", "*.xlsx"),("all files",
                                                            "*.*")]))
        entry4.xview_moveto(1)

    def close(self):
        pickle.dump(self.raw_input.get(),open("pref.dat", "wb"))
        self.root.destroy()

    def logprint(self,text):
        self.log.insert(END, text + '\n')

    def save(self):
        newWindow.destroy()

    def cancel(self):
        self.set_default()
        newWindow.destroy()

    def set_default(self):
        self.machine1_slop.set(self.machine1["slop"])
        self.machine2_slop.set(self.machine2["slop"])
        self.machine1_int.set(self.machine1["y-int"])
        self.machine2_int.set(self.machine2["y-int"])
        self.lower_cal.set(self.LOWER)
        self.upper_cal.set(self.UPPER)


    def setting(self):
        global newWindow

        newWindow = Toplevel(root)
        newWindow.title("Setting Window")
        row1 = ttk.Frame(newWindow)

        Label(row1, text="QS3",width=20).pack(side=LEFT, padx=5)
        Label(row1, text="QS5",width=20).pack(side=LEFT, padx=5)
        row1.pack(side=TOP, padx=5, pady=5)

        row2 = ttk.Frame(newWindow)
        Label(row2,text="Slop").pack(side=LEFT, padx=5)
        entry1 = ttk.Entry(row2,width=20,textvariable=self.machine1_slop)
        entry1.pack(side=LEFT, expand=YES, fill=X)
        Label(row2,text="Slop").pack(side=LEFT, padx=5)
        entry2 = ttk.Entry(row2,width=20,textvariable=self.machine2_slop)
        entry2.pack(side=LEFT, expand=YES, fill=X)
        row2.pack(side=TOP, padx=5, pady=5)

        row3 = ttk.Frame(newWindow)
        Label(row3,text="Y-int").pack(side=LEFT, padx=5)
        entry3 = ttk.Entry(row3,width=20,textvariable=self.machine1_int)
        entry3.pack(side=LEFT, expand=YES, fill=X)
        Label(row3,text="Y-int").pack(side=LEFT, padx=5)
        entry4 = ttk.Entry(row3,width=20,textvariable=self.machine2_int)
        entry4.pack(side=LEFT, expand=YES, fill=X)
        row3.pack(side=TOP, padx=5, pady=5)

        row5 = ttk.Frame(newWindow)
        Label(row5,text="Lower Cal").pack(side=LEFT, padx=5)
        entry5 = ttk.Entry(row5,width=20,textvariable=self.lower_cal)
        entry5.pack(side=LEFT, expand=YES, fill=X)
        Label(row5,text="Upper Cal").pack(side=LEFT, padx=5)
        entry6 = ttk.Entry(row5,width=20,textvariable=self.upper_cal)
        entry6.pack(side=LEFT, expand=YES, fill=X)
        row5.pack(side=TOP, padx=5, pady=5)

        row4 = ttk.Frame(newWindow)
        ttk.Button(row4, text="Save", command=self.save).pack(side=LEFT,padx=15)
        ttk.Button(row4, text="Cancel", command=self.cancel).pack(side=LEFT,padx=15)
        row4.pack(side=TOP, padx=5, pady=5)

        # self.set_default()


    def mapping(self):
        global  machine
        machine = {}
        try:
            inpath = Path(self.raw_input.get())
            df = pd.DataFrame()
            df = pd.read_excel(inpath, sheet_name=self.sheet_name)
        except:
            self.logprint("- Error reading input file!")
            messagebox.showerror('Error', 'Error reading input file!') # TODO show output file name
            return 1
        
        for index, row in df.iterrows():
            if row.str.contains("Instrument Type").any():
                machine_row = row
            if row.str.contains(self.starting_word).any():
                start_index = index
                break
        
        if start_index is not None:
            # Create a new DataFrame starting from the row containing the start_word
            new_df = pd.DataFrame()
            new_df = df.iloc[start_index:]
            new_df.reset_index(drop=True, inplace=True)
            new_df.columns = new_df.iloc[0]
            new_df = new_df.iloc[1:]
            new_df.reset_index(drop=True, inplace=True)
        else:
            print(f"Starting cell with the word '{self.starting_word}' not found in the DataFrame.")

        if machine_row is not None:
            machine_val = machine_row.iloc[1]  # Assuming "Instrument Name" is in the second column
            if "3" in machine_val:
                machine["slop"] = float(self.machine1_slop.get())
                machine["y-int"] = float(self.machine1_int.get())
            else:
                machine["slop"] = float(self.machine2_slop.get())
                machine["y-int"] = float(self.machine2_int.get())
            self.machine_str.set(f"{machine_val}")
            
        return new_df

    def findneg(self,df,wb,sheet):
        for row, value in enumerate(df[self.search_word],start=2):
            if isinstance(value, str) and len(value) > 0:
                if value.startswith("NEG"):
                    ct_mean = df.loc[row - 2, 'Ct Mean']
                    if not pd.isna(ct_mean) or ct_mean < 38:
                        cell = sheet.cell(row,column=2)
                        cell.fill = self.red_fill
                elif value.startswith("NTC"):
                    ct_mean = df.loc[row - 2, 'Ct Mean']
                    if not pd.isna(ct_mean):
                        cell = sheet.cell(row,column=2)
                        cell.fill = self.red_fill
                elif value.startswith("EXT"):
                    ct_mean = df.loc[row - 2, 'Ct Mean']
                    if not pd.isna(ct_mean) or ct_mean < 38:
                        cell = sheet.cell(row,column=2)
                        cell.fill = self.red_fill

        wb.save(self.ofname.get())

    def checkcal(self,df,wb,sheet):
        for row, value in enumerate(df[self.search_word],start=2):
            if isinstance(value, str) and len(value) > 0:
                if value.startswith("Cal"):
                    ct_mean = df.loc[row - 2, 'Ct Mean']
                    if ct_mean < float(self.lower_cal.get()) or ct_mean > float(self.upper_cal.get()) :
                        cell = sheet.cell(row,column=2)
                        cell.fill = self.red_fill
        wb.save(self.ofname.get())

    def checkmean_and_sd(self,df,wb, sheet):

        # Iterate through the 'Ct Mean' column and fill with red if > 31
        for row_number, ct_mean_value in enumerate(df['Ct Mean'], start=2):  # Start at row 2 (assuming header is in row 1)
            if ct_mean_value > 31:
                cell = sheet.cell(row_number,column=2)
                cell.fill = self.red_fill

        for row_number, ct_sd_value in enumerate(df['Ct SD'], start=2):  # Start at row 2 (assuming header is in row 1)
            if ct_sd_value > 0.2:
                cell = sheet.cell(row_number,column=3)
                cell.fill = self.red_fill
        # Save the Excel file with modified formatting
        wb.save(self.ofname.get())

    def checkresult(self,df):
        try:
            wb = openpyxl.load_workbook(self.ofname.get())

            # Access the sheet
            sheet = wb['Sheet1']
        except:
            self.logprint("cant open the file")
        self.checkmean_and_sd(df,wb,sheet)
        self.findneg(df,wb,sheet)
        self.checkcal(df,wb,sheet)

    def updatedropdown(self):
        global buttonrow3, drop2
        drop2['menu'].delete(0,'end')
        self.clicked2.set(sheet_names[0])
        for name in sheet_names:
            drop2['menu'].add_command(label=name, command=lambda name=name: self.clicked2.set(name))

    def loadconcerntrate(self):
        global sheet_names, wb
        sheet_names = []
        try:
            wb = openpyxl.load_workbook(self.master_input.get())
            sheet_names = wb.sheetnames

        except:
            self.logprint("cant open the file")
        if not sheet_names:
            self.logprint("No sheets found")
            return
        else:
            self.updatedropdown()
        self.logprint("- Loading Master sheets done !")

    def add_concentrate_dilution(self,df):
        global wb
        try:
            df1 = df
            sheet = wb[self.clicked2.get()]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Convert the list of lists to a pandas DataFrame
            df2 = pd.DataFrame(data)
            df2.columns = df2.iloc[0]
            df2 = df2[1:]
            # df2 = pd.read_excel(master_input.get(),clicked2.get())
            for index, row in df1.iterrows():   
                master_id = row[self.search_word]
                if not pd.isna(master_id):
                    matching_row_df2 = df2[df2['[Sample ID]'].astype(str).str.contains(str(master_id))]
                if not matching_row_df2.empty:
            # Extract the value from the matching row in df2
                    for indexs in matching_row_df2.index:
                        concern = matching_row_df2.loc[indexs, '[Final Concentrate Volume (mL)]']
                        dilution = matching_row_df2.loc[indexs, '[Dilution factor]']
                        df1.loc[index, "Concentrate Volume (mL)"] = concern
                        df1.loc[index,"Dilution Factor"] = dilution
        except:
            self.logprint("- Error add concentration !")
            return 1
        return df1

    def outputConstrain(self,sample):
        # make an empty output constrain
        try:
            # write to the file
            sample_df = pd.DataFrame()
            sample_df = pd.DataFrame({self.search_word: sample,
                                        "Dilution Factor": [np.nan] * len(sample),
                                        "Concentrate Volume (mL)": [np.nan] * len(sample)
                                        
            })
        except:
            self.logprint("- Error writing constrain file!")
            messagebox.showerror('Error', 'Error writing constrain file!') 
            return 1
        df = self.add_concentrate_dilution(sample_df)
        return df

    def output_to_Excel(self,data,name):
        try:

            data.to_excel(name,index=False)
            
        except:
            self.logprint("- Error writing output file!")
            messagebox.showerror('Error', 'Error writing output file!') 
            return 1
        
    def extractConstrain(self):
        df = pd.DataFrame()
        df = self.mapping()
        # find the unique Sample ID and extract the column
        data = []
        sample = []
        data = df[self.search_word].unique()
        for x in data:
            # filter the control sets
            if (len(str(x)) >= 9):
                sample.append(x)
        return sample

    def runConstrain(self):
            df = pd.DataFrame()
            df = self.extractConstrain()
            df1 = self.outputConstrain(df)
            return df1


    def calculatePMMoV(self,ct_mean,concentrate,dilution):
        global machine
        pmmov1 = 10 ** ((ct_mean-machine["y-int"])/machine["slop"])
        pmmov2 = pmmov1/5 * dilution
        pmmov3 = pmmov2 *80
        pmmov4 = pmmov3/0.2*concentrate
        return pmmov4

    def result(self):
        # read two excel
        df1 = self.mapping()
        df2 = self.runConstrain()

        # merge two excel
        df3 = df1[self.filter_word].merge(df2,on = self.search_word, how = "left")

        #filter the excel
        df3['Sample Name'] = df3['Sample Name'].str.strip()
        df3_filter= df3.drop_duplicates()
        df3_filter = df3_filter.reset_index(drop=True)
        df3_filter["Calibrator Ct Mean"] = None
        df3_filter["PMMoV (gc/100 mL Sewage)"] = None

        for index,row in df3_filter.iterrows():
            ct_mean = row["Ct Mean"]
            concentrate = row["Concentrate Volume (mL)"]
            dilution = row["Dilution Factor"]
            # print(type(concentrate), type(dilution))
            if (isinstance(dilution, Number) and isinstance(concentrate, Number)):
                if not (math.isnan(dilution) or math.isnan(concentrate)):
                    dilution = int(dilution)  # Convert 'dilution' to an integer
                
            else:
                if not isinstance(dilution, Number):
                    dilution = 0
                if not isinstance(concentrate, Number):
                    concentrate = 0
            pmmov = self.calculatePMMoV(ct_mean,concentrate,dilution)
            df3_filter.loc[index, "PMMoV (gc/100 mL Sewage)"] = "{:.2e}".format(pmmov)

        
        for row, value in enumerate(df3_filter[self.search_word],start=2):
            if isinstance(value, str) and len(value) > 0:
                if value.startswith("Cal"):
                    ct_mean = df3_filter.loc[row - 2, 'Ct Mean']
                    df3_filter["Calibrator Ct Mean"] = ct_mean
        self.output_to_Excel(df3_filter,self.ofname.get())
        self.checkresult(df3_filter)
        self.logprint("- Outputting result file done !")
        try:
            os.system(f'start excel "{self.ofname.get()}"')
        except Exception as e:
            print(f"An error occurred: {e}")


    def clear(self):
        global sheet_names
        self.raw_input.set("")
        self.constrain_input.set("")
        self.ofname.set("")
        self.constrain_output.set("")
        self.master_input.set("")
        sheet_names = []
        self.clicked2.set(' ')
        self.drop2['menu'].delete(0,END)
        self.machine_str.set("")
        self.log.delete(1.0, END) 
        
if __name__ == '__main__':
    root = Tk()
    app = PMMoVCalc(root)
    root.mainloop()
